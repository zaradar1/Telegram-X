"""Export indexed messages/files to CSV, JSON, or a standalone HTML report."""

import csv
import json
from html import escape
from typing import Any, Dict, List

from utils import format_size


def export_csv(rows: List[Dict[str, Any]], file_path: str) -> None:
    if not rows:
        with open(file_path, "w", newline="", encoding="utf-8") as f:
            f.write("")
        return
    fieldnames = list(rows[0].keys())
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def export_json(rows: List[Dict[str, Any]], file_path: str) -> None:
    with open(file_path, "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, default=str)


def export_html(rows: List[Dict[str, Any]], file_path: str, title: str = "Telegram Manager Export") -> None:
    columns = list(rows[0].keys()) if rows else []
    header_html = "".join(f"<th>{escape(str(col))}</th>" for col in columns)

    body_rows = []
    for row in rows:
        cells = []
        for col in columns:
            val = row.get(col)
            if col == "file_size" and isinstance(val, (int, float)):
                val = format_size(val)
            cells.append(f"<td>{escape(str(val)) if val is not None else ''}</td>")
        body_rows.append("<tr>" + "".join(cells) + "</tr>")

    html = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>{escape(title)}</title>
<style>
body {{ font-family: -apple-system, Segoe UI, Arial, sans-serif; margin: 2rem; }}
table {{ border-collapse: collapse; width: 100%; }}
th, td {{ border: 1px solid #ddd; padding: 6px 10px; text-align: left; font-size: 13px; }}
th {{ background: #f4f4f4; position: sticky; top: 0; }}
tr:nth-child(even) {{ background: #fafafa; }}
</style></head>
<body>
<h2>{escape(title)}</h2>
<p>{len(rows)} rows</p>
<table><thead><tr>{header_html}</tr></thead><tbody>
{''.join(body_rows)}
</tbody></table>
</body></html>"""

    with open(file_path, "w", encoding="utf-8") as f:
        f.write(html)


def export_excel(rows: List[Dict[str, Any]], file_path: str, sheet_name: str = "Export") -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name length limit

    columns = list(rows[0].keys()) if rows else []
    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.font = Font(bold=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col in enumerate(columns, start=1):
            val = row.get(col)
            if col == "file_size" and isinstance(val, (int, float)):
                val = format_size(val)
            elif isinstance(val, (dict, list)):
                val = str(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx, col in enumerate(columns, start=1):
        width = max(10, min(60, len(str(col)) + 4))
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(file_path)


EXPORTERS = {
    "csv": export_csv,
    "json": export_json,
    "html": export_html,
    "xlsx": export_excel,
}


def export(rows: List[Dict[str, Any]], file_path: str, fmt: str) -> None:
    fmt = fmt.lower().lstrip(".")
    if fmt not in EXPORTERS:
        raise ValueError(f"Unsupported export format: {fmt}")
    EXPORTERS[fmt](rows, file_path)
